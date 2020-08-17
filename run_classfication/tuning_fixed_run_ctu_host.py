"""
Script to tune the parameters for the LSH heuristic
"""

import argparse
import glob
import inspect
import os
import shutil
import subprocess
import sys

import numpy as np
# import graphviz as graphviz
import openpyxl
from IPython.display import Image, display

currentdir = os.path.dirname(os.path.abspath(inspect.getfile(inspect.currentframe())))
parentdir = os.path.dirname(currentdir)
sys.path.insert(0, parentdir)

from pypackage.flexfringe.eval import profiling_evaluation, get_selectivity_threshold, error_based_classification, \
    fingerprint_based_classification, get_evaluation_metrics, get_profiling_threshold
from trigram import train_trigram_model

selected_features = ['duration', 'protocol', 'total_bytes', 'src_bytes', 'time_diff', 'packets']
majority = False

features_dir = '_'.join(selected_features)


def flexfringe(*args, **kwargs):
    command = ["--help"]

    if (len(kwargs) > 1):
        command = []
        for key in kwargs:
            command += ["-" + key + "=" + kwargs[key]]

    proc = subprocess.Popen(["../flexfringe", ] + command + [args[0]], stdout=subprocess.PIPE)
    output = proc.stdout.read()
    for line in output.decode("utf-8").split('\n'):
        print(line)
        if line.startswith('found intermediate solution with'):
            states = int(line.split()[4])


    dfa_path = '/'.join(args[0].split('/')[:-1]) + '/'
    try:
        with open(dfa_path + "final.dot") as fh:
            # return fh.read()
            return states
    except FileNotFoundError:
        pass
    return "No output file was generated."


def show(data):
    if data == "":
        pass
    else:
        g = graphviz.Source(data, format="png")
        g.render()
        display(Image(g.render()))


def write_to_csv(results, klthreshold, states, case, lsh=True):
    workbook_path = "../tuning_results/ctu/{}/hosts".format(features_dir)
    if not os.path.exists(workbook_path):
        os.makedirs(workbook_path)

    if lsh:
        file_workbook = workbook_path + "/tablesize_1000_hfun_100_1SD_3LSHs_ctu_lsh_0.01_results_largestblue_0_{}_{}_{}.xlsx".format(
            window,
            stride,
            case)
    else:
        file_workbook = workbook_path + "/Trigram_results_{}_{}_{}.xlsx".format(window, stride, case)

    if os.path.exists(file_workbook):
        book = openpyxl.load_workbook(file_workbook)
    else:
        book = openpyxl.Workbook()
    if sheet_name in book.sheetnames:
        del book[sheet_name]
    sheet1 = book.create_sheet(sheet_name)

    sheet1.cell(1, 2).value = "TP"
    sheet1.cell(1, 3).value = "FP"
    sheet1.cell(1, 4).value = "TN"
    sheet1.cell(1, 5).value = "FN"
    sheet1.cell(1, 6).value = "Accuracy"
    sheet1.cell(1, 7).value = "Precision"
    sheet1.cell(1, 8).value = "Recall"
    sheet1.cell(1, 9).value = "Threshold"
    sheet1.cell(1, 10).value = "States"

    sheet1.cell(2, 1).value = 'Error-based'

    for index, result in enumerate(results):
        sheet1.cell(index + 2, 2).value = str(result[0])
        sheet1.cell(index + 2, 3).value = str(result[1])
        sheet1.cell(index + 2, 4).value = str(result[2])
        sheet1.cell(index + 2, 5).value = str(result[3])
        sheet1.cell(index + 2, 6).value = str(round(result[4], 2))
        sheet1.cell(index + 2, 7).value = str(round(result[5], 2))
        sheet1.cell(index + 2, 8).value = str(round(result[6], 2))
        sheet1.cell(index + 2, 9).value = str(states)
        sheet1.cell(index + 2, 10).value = str(klthreshold)
    book.save(file_workbook)


def train_model(malicious_traces_file, difference):
    output_dir = malicious_traces_file + '_dir/'
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    parser = argparse.ArgumentParser()
    parser.add_argument('--mode', '-M', help="batch or stream depending on the mode of operation", type=str,
                        default='batch')
    parser.add_argument('--heuristic_name', '-hn', help="The Name of the heuristic", type=str, default='lsh4')
    parser.add_argument('--data_name', '-d', help="Type of data for the heuristic", type=str, default='lsh4_data')
    parser.add_argument('--state_count', '-q', help="state_count", type=str, default='10')
    parser.add_argument('--symbol_count', '-y', help="symbol_count", type=str, default='10')
    parser.add_argument('--satdfabound', '-D', help="satdfabound", type=str, default='200')
    # parser.add_argument('--delta', '-D', help="delta", type=str, default='0.95')
    parser.add_argument('--epsilon', '-e', help="epsilon", type=str, default='0.3')
    parser.add_argument('--seed', '-s', help="seed", type=str, default='42')
    parser.add_argument('--testmerge', '-t', help="testmerge", type=str, default='0')
    parser.add_argument('--klthreshold', '-T', help="klthreshold", type=str, default='0.01')
    parser.add_argument('--largestblue', '-a', help="largestblue", type=str, default='1')
    parser.add_argument('--finalred', '-f', help="finalred", type=str, default='0')
    parser.add_argument('--tablesize', '-S', help="tablesize", type=str, default='1000')
    parser.add_argument('--numoftables', '-N', help="numoftables", type=str, default='5')
    parser.add_argument('--vectordimension', '-v', help="vectordimension", type=str,
                        default='100')  # exeprioments done with 50
    parser.add_argument('--binarycodebytes', '-c', help="binarycodebytes", type=str, default='100')
    parser.add_argument('--difference', '-i', help="difference", type=str, default='20')
    parser.add_argument('--width', '-W', help="bin width", type=str, default='20')
    parser.add_argument('--index', '-g', help="Index mode, you can choose 1(CAUCHY) or 2(GAUSSIAN)", type=str,
                        default='1')
    parser.add_argument('--output_dir', '-o', help="output-dir", type=str, default=output_dir)
    parser.add_argument('--train_file', '-trf', help="Train file", type=str,
                        default=malicious_traces_file)

    # parser.add_argument('--dot_file', '-df', help="Dot file", type=str,
    #                     default=dot_file)

    parser.add_argument('--test_file', '-tsf', help="Test file for SPICE", type=str,
                        default='data/PAutomaC-competition_sets/11.pautomac.test')

    parser.add_argument('--prefix_file', '-pf', help="Prefix file for SPICE", type=str,
                        default='data/SPiCe_Offline/prefixes/14.spice.prefix.public')

    parser.add_argument('--target_file', '-tf', help="Target file for SPICE", type=str,
                        default='data/SPiCe_Offline/targets/14.spice.target.public')

    args = parser.parse_args()

    states = flexfringe(args.train_file, M=args.mode, h=args.heuristic_name, d=args.data_name, q=args.state_count,
                        y=args.symbol_count, a=args.largestblue, f=args.finalred, D=args.satdfabound, e=args.epsilon,
                        s=args.seed, t=args.testmerge, T=args.klthreshold, o=args.output_dir,
                        i=difference, c=args.binarycodebytes, v=args.vectordimension, S=args.tablesize,
                        N=args.numoftables, W=args.width, g=args.index)
    return states, args.klthreshold


def evaluate_single_scenario(all_hosts, malicious_traces_file, configuration_traces_files, lsh=True):
    if lsh:
        with open(malicious_traces_file, 'r') as f:
            i = f.readline().split()[1]
        klthreshold, states = train_model(malicious_traces_file, str(int(i)))
        model = malicious_traces_file + '_dir/final.dot'
    else:
        model = train_trigram_model(malicious_traces_file)
        klthreshold, states = 0, 0
    all_hosts = [h for h in all_hosts if not h.endswith('_dir')]

    host_classification = {}
    host_classification['prof'] = {}
    host_classification['error'] = {}
    host_classification['finger'] = {}
    results = []

    threshold = get_selectivity_threshold(model, configuration_traces_files, malicious_traces_file)
    print("Error-based threshold: ", str(threshold))
    host_labels_err = []
    predicted_labels_err = []
    for host in all_hosts:
        if 'benign' in host:
            host_label = 0
        else:
            host_label = 1
        evaluation_result = error_based_classification(model, host,
                                                       malicious_traces_file, threshold)
        if evaluation_result == host_label:
            host_classification['error'][host] = True
        else:
            host_classification['error'][host] = False
        host_labels_err += [host_label]
        predicted_labels_err += [evaluation_result]
    print('\nError Based Classification')
    error_eval = get_evaluation_metrics(host_labels_err, predicted_labels_err)
    results += [list(error_eval)]

    write_to_csv(results, klthreshold, states, 'single', lsh=lsh)


def evaluate_multiple_scenarios(models, all_hosts, configuration_traces_files, malicious_traces_files, lsh=True):
    results = []

    host_labels_err = []
    predicted_labels_err = []
    threshold = []
    for i, model in enumerate(models):
        threshold += [get_selectivity_threshold(model, configuration_traces_files, malicious_traces_files[i])]
    for i, host in enumerate(all_hosts):
        malicious_found = False
        print('Error Based Classification: {} of {}'.format(i + 1, len(all_hosts)))
        if 'benign' in host:
            host_label = 0
        else:
            host_label = 1
        evaluation_results = []
        for i, model in enumerate(models):
            r = error_based_classification(model, host, malicious_traces_files[i], threshold[i])
            evaluation_results += [r]
            if r == 1:
                malicious_found = True
                break
        if malicious_found:
            evaluation_result = 1
        else:
            evaluation_result = 0

        host_labels_err += [host_label]
        predicted_labels_err += [evaluation_result]
    print('\nError Based Classification')
    error_eval = get_evaluation_metrics(host_labels_err, predicted_labels_err)
    results += [list(error_eval)]
    write_to_csv(results, 0, 0, 'multiple', lsh=lsh)


def run_single_scenario_evaluation(lsh=True):
    global sheet_name
    global window
    global stride
    global data_dir
    global scenario
    window = '20'
    stride = '10'
    rootdir = '../data/discretized_fixed{}_{}_final/ctu13_host/single/{}'.format(window, stride, features_dir)
    for subdir, dirs, files in os.walk(rootdir):
        if len(subdir.replace('\\', '/').split('/')) != 7:
            continue

        scenario = subdir.split('_')[-1]
        print("Scenario: ", scenario)
        sheet_name = scenario

        # data_dir = os.path.join(subdir, dir)
        data_dir = subdir
        all_hosts = glob.glob(data_dir + "/*")
        malicious_traces_files = [f for f in all_hosts if ('infected' in f and not f.endswith("_dir"))]
        benign_traces_files = [f for f in all_hosts if 'benign' in f]
        configuration_traces_files = [f for f in all_hosts if 'configuration' in f]
        malicious_traces_file = malicious_traces_files[0]
        all_hosts.remove(malicious_traces_file)
        for conf_host in configuration_traces_files:
            all_hosts.remove(conf_host)
        evaluate_single_scenario(all_hosts, malicious_traces_file, configuration_traces_files, lsh=lsh)


def run_multiple_scenarios_evaluation(lsh=True):
    global sheet_name
    global window
    global stride
    global data_dir
    global scenario
    window = '20'
    stride = '10'
    rootdir = '../data/discretized_fixed{}_{}_final/ctu13_host/multi/{}'.format(window, stride, features_dir)
    training_sets = [44, 45, 46, 48, 51, 52, 52, 54]
    evaluation_sets = [42, 43, 47, 49, 50]
    models = []
    given_malicious_trace = []
    # configuration_traces_files = []
    conf_hosts = glob.glob(rootdir + '/configuration_data/*')
    configuration_traces_files = [f for f in conf_hosts]
    for training_set in training_sets:
        for subdir, dirs, files in os.walk(rootdir + '/scenario_' + str(training_set)):
            if len(subdir.replace('\\', '/').split('/')) != 7:
                continue
            scenario = subdir.split('_')[-1]
            sheet_name = scenario

            data_dir = subdir
            dir_files = os.listdir(data_dir)
            for item in dir_files:
                if os.path.isdir(item):
                    shutil.rmtree(item)
            all_hosts = glob.glob(data_dir + "/*")
            malicious_traces_files = [f for f in all_hosts if ('infected' in f and os.path.isfile(f))]
            for malicious_traces_file in malicious_traces_files:
                if lsh:
                    with open(malicious_traces_file, 'r') as f:
                        i = f.readline().split()[1]
                        states, klthreshold = train_model(malicious_traces_file, str(int(i)))
                    model = malicious_traces_file + '_dir/final.dot'
                    if os.path.exists(model):
                        models += [model]
                        given_malicious_trace += [malicious_traces_file]
                else:
                    model = train_trigram_model(malicious_traces_file)
                    models += [model]
                    klthreshold, states = 0, 0
                    given_malicious_trace += [malicious_traces_file]

    for evaluation_set in evaluation_sets:
        for subdir, dirs, files in os.walk(rootdir + '/scenario_' + str(evaluation_set)):
            if len(subdir.replace('\\', '/').split('/')) != 7:
                continue
            scenario = subdir.split('_')[-1]
            sheet_name = scenario
            data_dir = subdir

            all_hosts = glob.glob(data_dir + "/*")
            evaluate_multiple_scenarios(models, all_hosts, configuration_traces_files, given_malicious_trace,
                                        lsh=lsh)


if __name__ == '__main__':
    run_single_scenario_evaluation(lsh=True)
    run_single_scenario_evaluation(lsh=False)
    run_multiple_scenarios_evaluation(lsh=True)
    run_multiple_scenarios_evaluation(lsh=False)
