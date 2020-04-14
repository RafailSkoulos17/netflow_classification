import os
import sys

import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import seaborn as sns


def plot_evaluation(file_workbook, tlt):
    if os.path.exists(file_workbook):
        book = openpyxl.load_workbook(file_workbook)
    else:
        print("Wrong file path")
        sys.exit()

    results = {}
    scenarios = [s for s in book.sheetnames if s != 'Sheet']
    for sc in scenarios:
        sheet1 = book[str(sc)]
        results[sc] = {}
        results[sc]['Profiling'] = [float(sheet1.cell(2, i).value) for i in range(2, 9)]
        results[sc]['Error-based'] = [float(sheet1.cell(3, i).value) for i in range(2, 9)]
        results[sc]['Fingerprinting'] = [float(sheet1.cell(4, i).value) for i in range(2, 9)]

    # TP	FP	TN	FN	Accuracy	Precision	Recall	Threshold	States
    x = np.arange(len(scenarios))
    metrics = ['Accuracy', 'Precision', 'Recall']
    for i in range(4, 7):
        prof = [results[sc]['Profiling'][i] for sc in scenarios]
        err = [results[sc]['Error-based'][i] for sc in scenarios]
        fing = [results[sc]['Fingerprinting'][i] for sc in scenarios]

        width = 0.2
        ax = plt.subplot()
        plt.tight_layout()
        ax.set_ylabel(metrics[i - 4])
        ax.set_xlabel('Scenario')
        ax.set_xticks(x)
        if not isinstance(scenarios, list):
            scenarios = scenarios.tolist()
        # ax.set_xticklabels(scenarios)
        ax.set_ylim([0.1, 1])
        plt.title(tlt)
        ax.bar(x - width, prof, width=width, color='b', align='center', label='Profiling')
        ax.bar(x, err, width=width, color='g', align='center', label='Error-based')
        ax.bar(x + width, fing, width=width, color='r', align='center', label='Fingerprinting')

        plt.grid(True, 'major', 'y', ls='--', lw=.5, c='k', alpha=.3)
        ax.legend()
        plt.tight_layout()
        plt.show()


def plot_confusion_matrix(file_workbook, tlt):
    if os.path.exists(file_workbook):
        book = openpyxl.load_workbook(file_workbook)
    else:
        print("Wrong file path")
        sys.exit()

    results = {}
    scenarios = [s for s in book.sheetnames if s != 'Sheet']
    for sc in scenarios:
        sheet1 = book[str(sc)]
        results[sc] = {}
        results[sc]['Profiling'] = [float(sheet1.cell(2, i).value) for i in range(2, 9)]
        results[sc]['Error-based'] = [float(sheet1.cell(3, i).value) for i in range(2, 9)]
        results[sc]['Fingerprinting'] = [float(sheet1.cell(4, i).value) for i in range(2, 9)]

        ax = plt.subplot()
        # TP	FP	TN	FN	Accuracy	Precision	Recall	Threshold	States
        # cm = tn, fp, fn, tp
        # tp=5, fp=6, tn=500, fn=2
        tp, fp, tn, fn = results[sc]['Error-based'][:4]
        cm = np.array([[fn, tp], [tn, fp]])
        # cm = np.asarray(cm).reshape(2,2)
        sns.heatmap(cm, annot=True, ax=ax)  # annot=True to annotate cells

        # labels, title and ticks
        ax.set_xlabel('Predicted labels')
        ax.set_ylabel('True labels')
        ax.set_title('{}: Confusion Matrix for {}'.format(tlt, sc.rstrip('.pcap_ISCX')))
        ax.xaxis.set_ticklabels(['Benign', 'Malicious'])
        ax.yaxis.set_ticklabels(['Malicious', 'Benign'])
        plt.show()


# scenarios = np.array([42, 43, 47, 49, 50])
# file_workbook = "data/bidirectional_results/trigram_results_100median_50median_multiple.xlsx"
# tlt = 'Multiple 3grams'
# plot_evaluation(file_workbook, scenarios, tlt)
# file_workbook = "data/bidirectional_results/kldivergence_results_100median_50median_multiple.xlsx"
# tlt = 'Multiple LSH'
# plot_evaluation(file_workbook, scenarios, tlt)

scenarios = np.arange(42, 55)
file_workbook = "results/ids_connection/ids_connection_0.1_lsh_results_1000median_500median_single.xlsx"
tlt = 'Single LSH'
plot_evaluation(file_workbook, tlt)
plot_confusion_matrix(file_workbook, tlt)

# file_workbook = "results/ids_connection/ids_connection_trigram_results_1000median_500median_single.xlsx"
# tlt = 'Single 3grams'
# plot_evaluation(file_workbook, tlt)


# scenarios = np.arange(42, 55)
# file_workbook = "data/bidirectional_results/new_trigram_results_1000median_500median_single.xlsx"
# tlt = 'Single 3grams'
# plot_evaluation(file_workbook, scenarios, tlt)

# scenarios = np.arange(42, 55)
# file_workbook = "data/bidirectional_results/new_lsh_results_1000median_500median_single.xlsx"
# tlt = 'Single LSH'
# plot_evaluation(file_workbook, scenarios, tlt)

# scenarios = np.arange(42, 55)
# file_workbook = "data/bidirectional_results/new_new_lsh_results_1000median_500median_single.xlsx"
# tlt = 'Single new LSH'
# plot_evaluation(file_workbook, scenarios, tlt)
# #
# scenarios = np.array([42, 43, 47, 49, 50])
# file_workbook = "results/ctu_connection/fresh_trigram_results_1000median_500median_multiple.xlsx"
# tlt = 'Multiple 3grams'
# plot_evaluation(file_workbook, scenarios, tlt)

# file_workbook = "data/bidirectional_results/fresh_lsh_results_1000median_500median_multiple.xlsx"
# tlt = 'Multiple LSH 0.05'
# plot_evaluation(file_workbook, scenarios, tlt)

# file_workbook = "results/ctu_connection/fresh_0.1_lsh_results_1000median_500median_multiple.xlsx"
# tlt = 'Multiple LSH 0.01'
# plot_evaluation(file_workbook, scenarios, tlt)

# rootdir = "data/ids/"
# train_sets = ['Thursday-WorkingHours-Morning-WebAttacks.pcap_ISCX', 'Tuesday-WorkingHours.pcap_ISCX',
#               'Thursday-WorkingHours-Afternoon-Infilteration.pcap_ISCX']
# train_sets = [rootdir + d for d in train_sets]
# scenarios = set(glob.glob(rootdir + '*')) - set(train_sets)
# scenarios -= {'data/ids/configuration_data'}
# scenarios = [x.split('/')[-1] for x in list(scenarios)]
#
# file_workbook = "results/ids/ids_lsh_results_1000median_500median_multiple.xlsx"
# tlt = 'Multiple LSH'
# plot_evaluation(file_workbook, scenarios, tlt)
#
# file_workbook = "results/ids/ids_trigram_results_1000median_500median_multiple.xlsx"
# tlt = 'Multiple 3grams'
# plot_evaluation(file_workbook, scenarios, tlt)
#

# scenarios = np.array([42, 43, 47, 49, 50])
# file_workbook = "results/ctu_connection/fresh_trigram_results_1000median_500median_multiple3.xlsx"
# tlt = 'Multiple2 3grams'
# plot_evaluation(file_workbook, scenarios, tlt)

# file_workbook = "data/bidirectional_results/fresh_lsh_results_1000median_500median_multiple.xlsx"
# tlt = 'Multiple LSH 0.05'
# plot_evaluation(file_workbook, scenarios, tlt)

# file_workbook = "results/ctu_connection/fresh_0.1_lsh_results_1000median_500median_multiple3.xlsx"
# tlt = 'Multiple3 LSH 0.01'
# plot_evaluation(file_workbook, scenarios, tlt)
#
# file_workbook = "results/ctu_connection/fresh_0.1_lsh_results_1000median_500median_multiple4.xlsx"
# tlt = 'Multiple4 LSH 0.01'
# plot_evaluation(file_workbook, scenarios, tlt)
